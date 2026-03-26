def generer_grille(fichier, nom_feuille):
    import openpyxl

    excel = openpyxl.load_workbook(fichier, data_only=True)

    feuille = excel[nom_feuille]

    padding_de_grille = '|&#8288 {: style="padding:0"}'

    grille = []

    # Écrire les entêtes du tableau de l'horaire
    entete = ""
    trait = ""
    for col in range(1, feuille.max_column + 1):
        entete += feuille.cell(row=1, column=col).value
        trait += "--"
        if col < feuille.max_column:
            entete += "|"
            trait += "|"
        else:
            entete += "\n"
            trait += "\n"

    grille.append(entete)
    grille.append(trait)

    for row in range(2, feuille.max_row + 1):
        ligne = ""
        has_colspan = False
        for col in range(1, feuille.max_column + 1):
            valeur = feuille.cell(row=row, column=col).value
            if valeur is None:
                valeur = ""
            else:
                valeur = str(valeur).replace("\n", "")
            ligne += valeur
            if col < feuille.max_column:
                if not has_colspan:
                    ligne += "|"
            else:
                if has_colspan:
                    ligne += padding_de_grille * (feuille.max_column - 1)
                ligne += "\n"
            if "colspan" in valeur:
                has_colspan = True

        

        grille.append(ligne)
    
    return grille

def generer_horaire():

    grille = generer_grille("template/horaire.xlsx", "groupe1")

    with open("./wiki/horaire-groupe1.md", "w") as f:
        # Écrire le titre de la page
        f.write("# Horaire du cours de support technique\n")
        f.writelines(grille)


def generer_horaire_eleve():

    grille = generer_grille("template/horaire_eleve.xlsx", "Sheet1")

    with open("./wiki/horaire-eleve.md", "w") as f:
        # Écrire le titre de la page
        f.write("# Horaire simulation de formation\n")
        f.writelines(grille)

def generer_billeterie():
    intro = '''
# Utilisation d'une billeterie  

Dans ce cours, nous utiliserons __Notion__ comme système de billeterie.  

## Comment y accéder?  

Chaque équipe utilisera une billeterie différente. 

Voici les liens pour y accéder :'''
    
    billeteries = generer_grille("template/billeterie.xlsx", "liste")

    with open("./wiki/billeterie.md", "w") as f:
        f.write(intro)
        f.write("\n\n")
        f.writelines(billeteries)

print("Générer la page de l'horaire")
generer_horaire()
print("Générer la page de l'horaire pour les élèves")
generer_horaire_eleve()
print("Générer la page de la billeterie")
generer_billeterie()